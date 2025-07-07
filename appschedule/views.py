from django_filters.rest_framework import DjangoFilterBackend
from django.shortcuts import get_object_or_404
from rest_framework.permissions import IsAuthenticatedOrReadOnly, DjangoModelPermissions, DjangoModelPermissionsOrAnonReadOnly, IsAuthenticated
from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.authentication import TokenAuthentication
from rest_framework.exceptions import ValidationError
from rest_framework.decorators import action, api_view, permission_classes
from django.db.models import Q, Count
from django.db import connection                 # OAHP
from django.db.models.functions import TruncWeek # OAHP
from appschedule.models import Event             # OAHP
from django.utils import timezone
from django.utils.timezone import now            # OAHP
from asgiref.sync import async_to_sync            # OAHP
from django.contrib.auth import get_user_model   # OAHP
from channels.layers import get_channel_layer     # OAHP
from rest_framework.views import APIView
from rest_framework.pagination import PageNumberPagination

from .models import Event, EventDraft, EventNote, EventChatMessage, Crew, AbsenceReason, EventChatReadStatus
from crewsapp.models import Category, Job
from .serializers import EventSerializer, EventDraftSerializer, EventNoteSerializer, EventChatMessageSerializer, AbsenceReasonSerializer
from .filters import EventDraftFilter

import base64
from django.utils.dateparse import parse_date
from django.template.loader import render_to_string
from collections import defaultdict
from datetime import timedelta, datetime
from weasyprint import (
    CSS,
    HTML,
)
from weasyprint.text.fonts import FontConfiguration

from django.http import HttpResponse
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, PatternFill
from appidentity.models import Identity

class EventViewSet(viewsets.ModelViewSet):
    """ Event ViewSet """
    queryset = EventDraft.objects.all()
    serializer_class = EventDraftSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = EventDraftFilter
    permission_classes = [DjangoModelPermissions]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(created_by=self.request.user)

    def get_queryset(self):
        # if self.action == 'events_public':
        #     return Event.objects.select_related('crew').all()
        return EventDraft.objects.select_related('crew').all()

    def _publish_draft(self, draft):
        # print('draft.event_id:: ', draft.event)
        if draft.event_id is not None:
            event = Event.objects.get(id=draft.event_id)
            print('draft.__dict__:: ', event.__dict__)
        else:
            event = Event()
        excluded_fields = ['event', 'updated_at', 'created_at', 'pk', 'id']
        for field in EventDraft._meta.get_fields():
            if field.concrete and field.name not in excluded_fields:
                try:
                    setattr(event, field.name, getattr(draft, field.name))
                except AttributeError:
                    # Handle cases where the Event model might not have the exact same field
                    print(f"Warning: Event model does not have field '{field.name}'")
        event.save()
        draft.delete()
        print('order published: ')

    def create(self, request, *args, **kwargs):
        to_publish = request.data.pop('_post', False)
        if to_publish and not request.user.has_perm('appschedule.add_event'):
            return Response({'message': 'You do not have permission to publish events'}, status=status.HTTP_403_FORBIDDEN)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)

        if to_publish:
            draft = self.queryset.get(pk=serializer.data['id'])
            self._publish_draft(draft)
            return Response({'message': 'Draft published and deleted'}, status=status.HTTP_201_CREATED,
                            headers=headers)
        else:
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        to_publish = request.data.pop('_post', False)
        if to_publish and not request.user.has_perm('appschedule.add_event'):
            return Response({'message': 'You do not have permission to publish events'}, status=status.HTTP_403_FORBIDDEN)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            instance._prefetched_objects_cache = {}

        if to_publish:
            self._publish_draft(instance)
            return Response({'message': 'Draft updated, published and deleted'}, status=status.HTTP_200_OK)
        else:
            return Response(serializer.data, status=status.HTTP_200_OK)

    def destroy(self, request, *args, **kwargs):
        deleted = request.query_params.get('deleted', False)
        if deleted:
            event = Event.objects.get(pk=kwargs['pk'])
            event.deleted = True
            event.save()
            return Response({'message': f'Event with ID {event.id} has been deleted'}, status=status.HTTP_200_OK)
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'])
    def publish_drafts(self, request):
        """
        Publica los eventos draft dentro del rango de fechas proporcionado.
        Recibe 'start_date' y 'end_date' en el cuerpo de la petición POST.
        """
        start_date_str = request.data.get('start_date')
        end_date_str = request.data.get('end_date')

        if not start_date_str or not end_date_str:
            return Response({'error': 'Start and end dates must be provided.'},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = timezone.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'The date format must be YYYY-MM-DD.'},
                            status=status.HTTP_400_BAD_REQUEST)

        events_to_publish = EventDraft.objects.filter(
            date__gte=start_date,
            end_dt__lt=end_date
        )
        for event in events_to_publish:
            self._publish_draft(event)
        return Response({}, status=status.HTTP_200_OK)


class EventsListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        start_at = self.request.query_params.get('start_at')
        end_at = self.request.query_params.get('end_at')
        
        if not start_at:
            raise ValidationError('start_at must be provided')
        if not end_at:
            raise ValidationError('end_at must be provided')
        q = Q()
        q &= (Q(date__lte=start_at,
                end_dt__gte=start_at) |  # Empieza antes o en el inicio y termina después o en el inicio
              Q(date__gte=start_at, date__lt=end_at) |  # Empieza dentro del rango
              Q(date__lte=start_at, end_dt__lt=end_at))  # Termina dentro del rango
        q &= (Q(date__lte=end_at, end_dt__gte=end_at) |  # Empieza antes o en el fin y termina después o en el fin
              Q(date__gte=start_at, date__lte=end_at) |  # Empieza dentro del rango
              Q(end_dt__gte=start_at, end_dt__lte=end_at))  # Termina dentro del rango

        events = Event.objects.select_related('crew').filter(q, deleted=False)
        
        # Excluir los eventos que ya tienen draft (solo para usuarios con permiso)
        if request.user.has_perm('appschedule.add_eventdraft'):
            exclude_list = EventDraft.objects.exclude(event__isnull=True).values_list('event_id', flat=True)
            if len(exclude_list) > 0:
                events = events.exclude(id__in=exclude_list)

        serializer_event = EventSerializer(events.distinct(), many=True)
        response = {
            'events': serializer_event.data,
        }
        
        # Agregar drafts si el usuario tiene permiso
        if request.user.has_perm('appschedule.add_eventdraft'):
            drafts = EventDraft.objects.select_related('crew').filter(q).distinct()
            serializer_draft = EventDraftSerializer(drafts, many=True)
            response['drafts'] = serializer_draft.data
            
        # Agregar resumen por categoría del crew
        category_counts = (
            events
            .filter(is_absence=False)
            .values('crew__category__name')
            .annotate(total=Count('id'))
        )
        response['category_totals'] = list(category_counts)
            
        return Response(response, status=status.HTTP_200_OK)


class EventNoteViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = EventNoteSerializer
    lookup_field = 'event_id'

    def retrieve(self, request, event_id=None):
        try:
            event = get_object_or_404(Event, pk=event_id)
            try:
                note = EventNote.objects.get(event=event)
                serializer = self.serializer_class(note)
                return Response(serializer.data)
            except EventNote.DoesNotExist:
                return Response({'notes': ''}, status=status.HTTP_200_OK)  # Or 404 if you prefer
        except Event.DoesNotExist:
            return Response({'error': f'Event with ID {event_id} not found.'}, status=status.HTTP_404_NOT_FOUND)

    def create(self, request, event_id=None):
        try:
            get_object_or_404(Event, pk=event_id)
            event_note = EventNote.objects.filter(event=event_id).first()
            if event_note:
                serializer = self.serializer_class(event_note, data=request.data, context={'request': request})
            else:
                serializer = self.serializer_class(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


User = get_user_model()

class EventChatViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = EventChatMessageSerializer

    def list(self, request, event_id=None):
        if not event_id:
            return Response({"error": "Se debe proporcionar el ID del evento."}, status=400)

        event = get_object_or_404(Event, pk=event_id)
        queryset = EventChatMessage.objects.filter(event=event).order_by('timestamp')
        serializer = self.serializer_class(queryset, many=True)

        # Marcar como leído
        EventChatReadStatus.objects.update_or_create(
            event=event,
            user=request.user,
            defaults={'last_read': timezone.now()}
        )

        return Response(serializer.data)

    def create(self, request, event_id=None):
        if not event_id:
            return Response({"error": "Se debe proporcionar el ID del evento."}, status=400)

        event = get_object_or_404(Event, pk=event_id)
        serializer = self.serializer_class(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        message = serializer.save(event=event)

        # Notificar a otros usuarios con WebSocket
        channel_layer = get_channel_layer()
        for target_user in User.objects.exclude(id=request.user.id):
            last_read = EventChatReadStatus.objects.filter(event=event, user=target_user).first()
            if last_read:
                count = EventChatMessage.objects.filter(event=event, timestamp__gt=last_read.last_read).count()
            else:
                count = EventChatMessage.objects.filter(event=event).count()

            group_name = f"user_{target_user.id}_unread"
            async_to_sync(channel_layer.group_send)(
                group_name,
                {
                    "type": "unread.updated",
                    "event_id": event.id,
                    "count": count,
                    "user_id": request.user.id,
                }
            )

        return Response(serializer.data, status=201)

@permission_classes([IsAuthenticated])    
@api_view(['GET'])
def download_schedule_pdf(request):
    start_at = request.GET.get('start_at')
    end_at = request.GET.get('end_at')
    # print(f"🗕️ Dates: {start_at} - {end_at}")

    if not start_at or not end_at:
        return Response({'error': 'start_at and end_at are required'}, status=400)

    start_date = parse_date(start_at)
    end_date = parse_date(end_at)

    # end_date se toma como tope, no inclusive
    days = []
    current = start_date
    while current < end_date:
        days.append(current)
        current += timedelta(days=1)

    # print("🗖 Days to render:")
    for d in days:
        d.strftime('%A, %b %d')

    # Consulta eventos cruzados en el rango
    q = Q()
    q &= (Q(date__lte=start_date, end_dt__gte=start_date) |
          Q(date__gte=start_date, date__lt=end_date) |
          Q(date__lte=start_date, end_dt__lt=end_date))
    q &= (Q(date__lte=end_date, end_dt__gte=end_date) |
          Q(date__gte=start_date, date__lte=end_date) |
          Q(end_dt__gte=start_date, end_dt__lte=end_date))

    events = Event.objects.select_related('crew', 'crew__category').filter(q, deleted=False)

    # Todos los crews activos agrupados por categoría
    all_crews = Crew.objects.select_related('category').filter(status=True)
    categorized_events = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

    for crew in all_crews:
        if crew.category:
            categorized_events[crew.category.name][crew.name]  # inicializar aunque no tenga eventos

    for event in events:
        if not event.crew or not event.crew.category:
            continue

        crew_name = event.crew.name
        category_name = event.crew.category.name

        # Expande cualquier evento, marcando inicio y fin
        current = event.date
        while current < event.end_dt:
            if start_date <= current <= end_date:
                if event.is_absence and event.absence_reason:
                    absence_text = f"🛑 Absence: {event.absence_reason.name}"
                    if event.description:
                        absence_text += f" – {event.description}"
                    categorized_events[category_name][crew_name][current].append(absence_text)
                else:
                    if current == event.date:
                        categorized_events[category_name][crew_name][current].append(event)
                    else:
                        categorized_events[category_name][crew_name][current].append("Finishing up work")
            current += timedelta(days=1)

    categorized_events_clean = {
        cat: {
            crew: dict(days)
            for crew, days in crews.items()
        }
        for cat, crews in categorized_events.items()
    }

    domain = request.get_host()
    
    # if 'phoenixelectricandair' in domain:
    #     tenant_logo = 'media/tenant_logos/Logo-phoenix-w.png'
    # elif '192.168.0.248:8000' in domain or 'division16llc' in domain:
    #     tenant_logo = 'media/tenant_logos/Logo-division-w.png'
    # else:
    #     tenant_logo = 'media/tenant_logos/default-logo.png'
    #     if identity and identity.logo:
    #         tenant_logo = identity.logo.path 

    # Obtener el logo personalizado del usuario si existe
    identity = Identity.objects.filter(user=request.user).first()
    
    if identity and identity.logo:
        logo_identity = identity.logo.url
    else:
       logo_identity = '/media/tenant_logos/default-logo.png'

    logo_url = request.build_absolute_uri(logo_identity)

    context = {
        'categorized_events': categorized_events_clean,
        'date_range': f"{start_date.strftime('%b %d')} – {(end_date - timedelta(days=1)).strftime('%b %d, %Y')}",
        'days': days,
        'logo_url': logo_url
    }

    html = render_to_string('schedule_pdf.html', context)
    font_config = FontConfiguration()
    pdf_file = HTML(string=html).write_pdf(font_config=font_config)

    return Response({
        'file': base64.b64encode(pdf_file),
        'filename': f'schedule_{start_at}_to_{end_at}.pdf',
        'file_type': 'application/pdf'
    }, status=200)


class MyEventsView(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination
    pagination_class.page_size = 20

    def get(self, request, format=None):
        search = self.request.query_params.get('search')
        today = timezone.now().date()

        q = Q(deleted=False)

        # Verifica si el usuario no es superusuario
        if not request.user.is_superuser:
            jobs = Job.objects.filter(crews__members=request.user).values_list('id', flat=True)
            if jobs.exists():
                q &= Q(job__in=jobs)
                print(f"[DEBUG] {request.user.username} belongs to crews with job IDs: {list(jobs)}")
            else:
                print(f"[DEBUG] {request.user.username} does not belong to any crew — all events will be shown")

        # Búsqueda por palabra clave
        if search:
            q &= (
                Q(title__icontains=search) |
                Q(description__icontains=search) |
                Q(crew__name__icontains=search)
            )

        queryset = Event.objects.select_related('crew').filter(q, deleted=False).order_by('-date', 'crew__name').distinct()

        # Paginación
        paginator = self.pagination_class()
        page = paginator.paginate_queryset(queryset, request, view=self)
        serializer = EventSerializer(page, many=True)
        return paginator.get_paginated_response(serializer.data)
    
    
@permission_classes([IsAuthenticated])
@api_view(['GET'])
def export_schedule_excel(request):
    start_at = request.GET.get("start_at")
    end_at = request.GET.get("end_at")

    try:
        start_date = datetime.strptime(start_at, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_at, "%Y-%m-%d").date()
    except:
        return HttpResponse("Invalid dates", status=400)

    # Filtro robusto para eventos cruzados
    q = Q()
    q &= (Q(date__lte=start_date, end_dt__gte=start_date) |
          Q(date__gte=start_date, date__lte=end_date))
    q &= (Q(date__lte=end_date, end_dt__gte=end_date) |
          Q(end_dt__gte=start_date, end_dt__lte=end_date))

    events = Event.objects.filter(q, deleted=False).select_related("crew", "crew__category").order_by("crew__name")

    # Construir días
    days = [start_date + timedelta(days=i) for i in range((end_date - start_date).days)]

    # Agrupar eventos por categoría, crew y día (ajustando end_dt - 1 día)
    categorized_events = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    
    for event in events:
        category = event.crew.category.name if hasattr(event.crew, "category") and event.crew.category else "Uncategorized"
        crew_name = event.crew.name
        adjusted_end = event.end_dt - timedelta(days=1)  # Aquí - 1 dia
        event_range = [event.date + timedelta(days=i) for i in range((adjusted_end - event.date).days + 1)]
        for day in days:
            if day in event_range:
                info = f"{event.title}"
                if event.extended_service:
                    info += " ⚡Ext"
                if event.description:
                    info += f" - {event.description}"
                categorized_events[category][crew_name][day].append(info)

    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # borrar la hoja vacía por defecto

    for category, crews in categorized_events.items():
        ws = wb.create_sheet(title=category[:31])  # máximo 31 caracteres
        header = ["Crew"] + [d.strftime("%Y-%m-%d") for d in days]
        ws.append(header)

        # Aplicar negrita al header
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")

        for crew_name, events_by_day in crews.items():
            row = [crew_name]
            for d in days:
                daily_events = events_by_day[d]
                cell_text = "\n".join(daily_events) if daily_events else ""
                row.append(cell_text)
            ws.append(row)

        # Ajustar estilos
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        for col in range(1, len(header) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    # Devolver el archivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"schedule_{start_date}_to_{(end_date - timedelta(days=1))}.xlsx"
    print(filename)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

class AbsenceReasonViewSet(viewsets.ModelViewSet):
    queryset = AbsenceReason.objects.filter(is_active=True)
    serializer_class = AbsenceReasonSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    
    
class WeeklySupervisorStatsChartView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Leer parámetros del querystring
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        category = request.query_params.get('category')

        # Si no hay fechas en los query params, usar 12 semanas atrás hasta hoy
        today = datetime.today().date()
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date() if start_date_str else today - timedelta(weeks=12)
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date() if end_date_str else today
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)

        # Armamos filtro dinámico y lista de parámetros
        category_filter = ""
        params = [start_date, end_date]

        if category:
            category_filter = "AND cc.name = %s"
            params.append(category)

        with connection.cursor() as cursor:
            cursor.execute(f"""
                SELECT
                    DATE_SUB(e.date, INTERVAL WEEKDAY(e.date) DAY) AS week,
                    au.username AS supervisor,
                    COUNT(DISTINCT e.id) AS total_events
                FROM appschedule_event e
                JOIN crewsapp_crew c ON e.crew_id = c.id
                JOIN crewsapp_crew_jobs crj ON e.job_id = crj.job_id
                JOIN crewsapp_crew_members crm ON crj.crew_id = crm.crew_id
                JOIN auth_user au ON crm.user_id = au.id
                JOIN crewsapp_category cc ON c.category_id = cc.id
                WHERE e.deleted = FALSE
                  AND e.is_absence = FALSE
                  AND e.date BETWEEN %s AND %s
                  {category_filter}
                GROUP BY week, au.username
                ORDER BY week, au.username
            """, params)

            rows = cursor.fetchall()

        # Procesamos los datos para Chart.js
        data_map = {}
        labels_set = set()

        for week, supervisor, total in rows:
            week_str = week.strftime('%m-%d-%Y')
            labels_set.add(week_str)
            if supervisor not in data_map:
                data_map[supervisor] = {}
            data_map[supervisor][week_str] = total

        labels = sorted(labels_set)
        datasets = []

        for supervisor, week_data in data_map.items():
            dataset = {
                "label": supervisor,
                "data": [week_data.get(week, 0) for week in labels]
            }
            datasets.append(dataset)

        return Response({
            "labels": labels,
            "datasets": datasets
        })
        
        
class WeeklySupervisorStatsExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Rango de semanas dinámico
        range_weeks = int(request.query_params.get('weeks', 16))  # por defecto 16
        today = datetime.today().date()

        # Fechas
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        category = request.query_params.get('category')

        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date() if start_date_str else today - timedelta(weeks=range_weeks)
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date() if end_date_str else today
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)

        # SQL y lógica de stats
        category_filter = ""
        params = [start_date, end_date]

        if category:
            category_filter = "AND cc.name = %s"
            params.append(category)

        with connection.cursor() as cursor:
            cursor.execute(f"""
                SELECT
                    DATE_SUB(e.date, INTERVAL WEEKDAY(e.date) DAY) AS week,
                    au.username AS supervisor,
                    COUNT(DISTINCT e.id) AS total_events
                FROM appschedule_event e
                JOIN crewsapp_crew c ON e.crew_id = c.id
                JOIN crewsapp_crew_jobs crj ON e.job_id = crj.job_id
                JOIN crewsapp_crew_members crm ON crj.crew_id = crm.crew_id
                JOIN auth_user au ON crm.user_id = au.id
                JOIN crewsapp_category cc ON c.category_id = cc.id
                WHERE e.deleted = FALSE
                  AND e.is_absence = FALSE
                  AND e.date BETWEEN %s AND %s
                  {category_filter}
                GROUP BY week, au.username
                ORDER BY week, au.username
            """, params)

            rows = cursor.fetchall()

        # Prepara data
        data_map = {}
        weeks_set = set()

        for week, supervisor, total in rows:
            week_str = week.strftime('%Y-%m-%d')
            weeks_set.add(week_str)
            data_map.setdefault(supervisor, {})[week_str] = total

        sorted_weeks = sorted(weeks_set)
        supervisors = sorted(data_map.keys())

        # Crea el Excel

        wb = Workbook()
        ws = wb.active
        ws.title = "Supervisor Stats"

        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Headers
        headers = ["Week"] + supervisors
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold
            cell.alignment = center
            cell.border = border

        # Data
        for week in sorted_weeks:
            row = [week]
            for supervisor in supervisors:
                row.append(data_map[supervisor].get(week, 0))
            ws.append(row)

        # Ajusta ancho
        for col in ws.columns:
            length = max(len(str(cell.value or "")) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = length + 2

        # Descarga
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"supervisor_report_{today}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def unread_chat_counts(request):
    user = request.user
    unread_counts = {}

    events = Event.objects.all()
    for event in events:
        last_read = EventChatReadStatus.objects.filter(event=event, user=user).first()
        if last_read:
            count = EventChatMessage.objects.filter(event=event, timestamp__gt=last_read.last_read).count()
        else:
            count = EventChatMessage.objects.filter(event=event).count()
        if count > 0:
            unread_counts[event.id] = count

    return Response(unread_counts)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_chat_read(request, event_id):
    user = request.user
    event = get_object_or_404(Event, pk=event_id)

    obj, created = EventChatReadStatus.objects.get_or_create(event=event, user=user)
    obj.last_read = now()
    obj.save()

    return Response({"status": "read updated"})